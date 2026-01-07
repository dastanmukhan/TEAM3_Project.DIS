import pandas as pd
import plotly.express as px
from sqlalchemy import create_engine, text
import plotly.io as pio
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

pio.renderers.default = 'browser'

engine = create_engine("postgresql+psycopg2://postgres:123456789@localhost:5432/hospital_db")

query = text("""
SELECT 
    b.bill_date as date, 
    b.payment_status,
    COALESCE(SUM(b.amount), 0) as total_amount
FROM billing b
WHERE b.bill_date LIKE '2023%'
GROUP BY b.bill_date, b.payment_status
ORDER BY b.bill_date, b.payment_status;
""")

df = pd.read_sql(query, engine)

if df.empty:
    print("Error: DataFrame is empty. Check database tables or query.")
else:
    print("DataFrame shape:", df.shape)
    print("DataFrame head:\n", df.head())

    df['date'] = pd.to_datetime(df['date'])

    fig = px.line(df, 
                  x="date", 
                  y="total_amount",
                  color="payment_status",
                  title="Daily Total Billed Amounts by Payment Status in 2023",
                  labels={"total_amount": "Total Amount ($)", "date": "Date"})

    fig.update_layout(
        xaxis=dict(
            rangeslider=dict(visible=True),
            type="date"
        ),
        yaxis_range=[0, df['total_amount'].max() * 1.2]
    )

    fig.show()

def export_to_excel(dataframes, filename):
    path = os.path.join("./exports/", filename)
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        for sheet_name, df in dataframes.items():
            safe_sheet = sheet_name[:31] if len(sheet_name) > 31 else sheet_name
            df.to_excel(writer, sheet_name=safe_sheet, index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        max_col = ws.max_column
        max_row = ws.max_row
        if max_row < 2:
            continue
        for col_idx in range(1, max_col + 1):
            sample = ws.cell(row=2, column=col_idx).value
            if isinstance(sample, (int, float)):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                rng = f"{col_letter}2:{col_letter}{max_row}"
                rule = ColorScaleRule(start_type='min', start_color='FF0000',
                                      mid_type='percentile', mid_value=50, mid_color='FFFF00',
                                      end_type='max', end_color='00FF00')
                ws.conditional_formatting.add(rng, rule)
                ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=[f"MAX({rng})"], stopIfTrue=True,
                                                              fill=PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')))
                ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=[f"MIN({rng})"], stopIfTrue=True,
                                                              fill=PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')))
            elif isinstance(sample, pd.Timestamp):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                for row in range(2, max_row + 1):
                    ws.cell(row=row, column=col_idx).number_format = 'YYYY-MM-DD'
                ws.column_dimensions[col_letter].width = 15

    wb.save(path)
    print(f"+ Created file {os.path.relpath(path, os.getcwd())}, sheets: {len(wb.sheetnames)}, rows: {wb.worksheets[0].max_row}")

dataframes = {
    "Billing_Summary": df
}
export_dir = "./exports/"
if not os.path.exists(export_dir):
    os.makedirs(export_dir)
    print(f"Created directory: {export_dir}")
filename = "billing_report.xlsx"
export_to_excel(dataframes, filename)